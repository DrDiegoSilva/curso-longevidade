"""Gera as 7 ferramentas baixáveis da trilha de empreendedorismo médico.

Arquivos ESTÁTICOS: são comitados em `seed/trilha/ferramentas/` e o app só
serve o download (`trilha.caminho_ferramenta`). Por isso pode usar openpyxl —
a regra de "só stdlib" vale para o container, não para o que geramos aqui.

Rode de novo sempre que o texto de uma das peças (01, 03, 04, 05, 07, 09, 12)
mudar, pra manter a ferramenta em sincronia com o que a peça manda o médico
fazer. Uso:

    python3 scripts/gerar_ferramentas_trilha.py

Padrão visual e de código: mesmo de `planilha-precificacao.xlsx` (a
referência já aprovada pelo dono do produto) — células amarelas = o médico
preenche, verdes = calculadas por fórmula (nunca valor fixo), seções com
faixa cinza, formatação condicional onde ajuda a enxergar problema.

Os números de exemplo que nascem preenchidos nas planilhas são os do cânone
da trilha (consultório de emagrecimento: custo da hora R$ 408, plano de 3
meses = 3,5 h, preço de tabela R$ 2.250) — ver `CANON.md` da sessão que
criou este script. Não inventar estatística fora desse cânone.
"""
import os
import shutil

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# --------------------------------------------------------------------- estilo
BRL = 'R$ #,##0.00'
PCT = '0"%"'          # número cru (40) exibido como "40%" — entradas de %
PCT_CALC = '0.0%'     # fração de fórmula (0.10) exibida como "10,0%"

TINTA = "1B211E"
VERDE = "1F6B58"
CINZA = "7C857F"
AMARELO = PatternFill("solid", fgColor="FFF4CE")
RESULTADO = PatternFill("solid", fgColor="E4EFEA")
TITULO_FILL = PatternFill("solid", fgColor="1F6B58")
SECAO_FILL = PatternFill("solid", fgColor="EDEFEA")
NEGATIVO_FILL = PatternFill("solid", fgColor="F6E2DE")
NEGATIVO_FONT = Font(color="9C3226", bold=True)

fina = Side(style="thin", color="C9CEC6")
BORDA = Border(left=fina, right=fina, top=fina, bottom=fina)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEST_DIR = os.path.join(REPO_ROOT, "seed", "trilha", "ferramentas")
# fonte da planilha de precificação já aprovada nesta sessão — se não existir
# mais (scratchpad é efêmero), o script regera pelo código embutido abaixo,
# que é a MESMA lógica, só com o destino trocado.
FONTE_PRECIFICACAO_APROVADA = (
    "/private/tmp/claude-501/-Users-diegosilva-dev-curso-longevidade"
    "--claude-worktrees-aviso-venda-whatsapp/da9b64ed-5a7e-4bb7-809c-557e935113bf"
    "/scratchpad/planilha-precificacao.xlsx"
)


# ------------------------------------------------------------- helpers base
def rotulo(ws, linha, texto, coluna=1, negrito=False, cor=TINTA, tamanho=11):
    c = ws.cell(row=linha, column=coluna, value=texto)
    c.font = Font(name="Calibri", size=tamanho, bold=negrito, color=cor)
    return c


def entrada(ws, linha, texto, valor, coluna_valor=2, fmt=BRL, dica="", coluna_dica=3):
    """Linha rótulo (coluna A) + valor amarelo editável (coluna B, por padrão)."""
    rotulo(ws, linha, texto)
    c = ws.cell(row=linha, column=coluna_valor, value=valor)
    c.fill, c.border, c.number_format = AMARELO, BORDA, fmt
    c.font = Font(name="Calibri", size=11)
    if dica:
        ws.cell(row=linha, column=coluna_dica, value=dica).font = Font(
            name="Calibri", size=9, italic=True, color=CINZA)
    return c


def calculado(ws, linha, texto, formula, coluna_valor=2, fmt=BRL, forte=False,
              dica="", coluna_dica=3):
    """Linha rótulo + valor verde calculado por fórmula (coluna B, por padrão)."""
    rotulo(ws, linha, texto, negrito=forte, cor=VERDE if forte else TINTA)
    c = ws.cell(row=linha, column=coluna_valor, value=formula)
    c.fill, c.border, c.number_format = RESULTADO, BORDA, fmt
    c.font = Font(name="Calibri", size=13 if forte else 11, bold=forte,
                  color=VERDE if forte else TINTA)
    if dica:
        ws.cell(row=linha, column=coluna_dica, value=dica).font = Font(
            name="Calibri", size=9, italic=True, color=CINZA)
    return c


def secao(ws, linha, texto, ncols=5):
    for col in range(1, ncols + 1):
        c = ws.cell(row=linha, column=col)
        c.fill = SECAO_FILL
        if col == 1:
            c.value = texto
            c.font = Font(name="Calibri", size=11, bold=True, color=VERDE)


def cabecalho_sheet(ws, nome, subtitulo, ncols=5):
    ultima = get_column_letter(ncols)
    ws.merge_cells(f"A1:{ultima}1")
    t = ws["A1"]
    t.value = nome
    t.fill = TITULO_FILL
    t.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{ultima}2")
    ws["A2"].value = subtitulo
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="4A534E")
    ws.row_dimensions[2].height = 22


def nota_rodape(ws, linha, texto, ncols=5):
    ultima = get_column_letter(ncols)
    ws.merge_cells(f"A{linha}:{ultima}{linha}")
    c = ws[f"A{linha}"]
    c.value = texto
    c.font = Font(name="Calibri", size=10, italic=True, color="4A534E")


def cabecalho_tabela(ws, linha, textos, col_inicio=1):
    for i, texto in enumerate(textos):
        c = ws.cell(row=linha, column=col_inicio + i, value=texto)
        c.font = Font(name="Calibri", size=9, bold=True, color=CINZA)
        c.border = BORDA
        c.alignment = Alignment(wrap_text=True, vertical="center")


# ----------------------------------------------------- helpers de tabela larga
def cel_amarela(ws, linha, coluna, valor=None, fmt=None, wrap=False):
    c = ws.cell(row=linha, column=coluna, value=valor)
    c.fill, c.border = AMARELO, BORDA
    if fmt:
        c.number_format = fmt
    c.font = Font(name="Calibri", size=11)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def cel_fixa(ws, linha, coluna, valor, negrito=False, wrap=False, cor=TINTA):
    c = ws.cell(row=linha, column=coluna, value=valor)
    c.border = BORDA
    c.font = Font(name="Calibri", size=11, bold=negrito, color=cor)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def cel_verde(ws, linha, coluna, formula, fmt=None, forte=False):
    c = ws.cell(row=linha, column=coluna, value=formula)
    c.fill, c.border = RESULTADO, BORDA
    if fmt:
        c.number_format = fmt
    c.font = Font(name="Calibri", size=13 if forte else 11, bold=forte,
                  color=VERDE if forte else TINTA)
    return c


def cel_neutra(ws, linha, coluna, valor="—"):
    c = ws.cell(row=linha, column=coluna, value=valor)
    c.border = BORDA
    c.font = Font(name="Calibri", size=10, italic=True, color=CINZA)
    c.alignment = Alignment(horizontal="center")
    return c


def larguras(ws, mapa):
    for col, w in mapa.items():
        ws.column_dimensions[col].width = w


# =============================================================== peça 01
def construir_custo_hora():
    """Custos fixos + pró-labore + horas atendidas -> custo da hora.
    Com os valores do cânone, tem que dar ~R$ 408."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Custo da hora"
    ws.sheet_view.showGridLines = False

    cabecalho_sheet(ws, "CUSTO REAL DA SUA HORA",
                     "Preencha só as células amarelas. As verdes se calculam sozinhas. "
                     "Semana 1 da trilha.")

    secao(ws, 4, "1 · QUANTO SAI DO SEU CONSULTÓRIO POR MÊS")
    entrada(ws, 5, "Aluguel e condomínio", 4500)
    entrada(ws, 6, "Secretária (com encargos)", 2800)
    entrada(ws, 7, "Prontuário e software", 300)
    entrada(ws, 8, "Contador", 600)
    entrada(ws, 9, "Energia, internet, limpeza", 700)
    entrada(ws, 10, "Marketing", 1200)
    entrada(ws, 11, "Outros", 0)
    entrada(ws, 12, "Seu pró-labore", 25000, dica="quanto VOCÊ precisa tirar por mês")
    calculado(ws, 13, "Total a cobrir por mês", "=SUM(B5:B12)", forte=True)

    secao(ws, 15, "2 · QUANTAS HORAS VOCÊ REALMENTE ATENDE")
    entrada(ws, 16, "Horas atendidas por semana", 20, fmt='0.0',
            dica="só as horas com paciente na sala, não a agenda cheia")
    entrada(ws, 17, "Semanas por mês", 4.3, fmt='0.0')
    calculado(ws, 18, "Horas atendidas por mês", "=B16*B17", fmt='0.0')
    calculado(ws, 19, "CUSTO DA SUA HORA", "=IF(B18=0,0,B13/B18)", forte=True,
              dica="abaixo disso você paga para atender, não é pago para atender")

    nota_rodape(ws, 21, "Esse número reaparece nas próximas 11 peças desta trilha — "
                        "guarde ele num lugar que você não vai perder.")

    larguras(ws, {"A": 38, "B": 16, "C": 34, "D": 12, "E": 12})
    return wb


# =============================================================== peça 03
def construir_mapa_de_linha():
    """Compara 3 linhas candidatas nos 4 critérios da peça 3 e aponta a
    vencedora por fórmula (nota total mais alta)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de linha"
    ws.sheet_view.showGridLines = False

    cabecalho_sheet(ws, "MAPA DE LINHA — QUAL VIRA SUA ESPECIALIDADE",
                     "Preencha as células amarelas com números reais da sua agenda dos "
                     "últimos 30 dias — não estime de memória. Semana 3 da trilha.",
                     ncols=5)

    secao(ws, 4, "1 · AS TRÊS LINHAS CANDIDATAS", ncols=5)
    cabecalho_tabela(ws, 5, ["Critério", "Linha 1", "Linha 2", "Linha 3"])
    ws.row_dimensions[5].height = 26

    cel_fixa(ws, 6, 1, "Nome da linha")
    cel_amarela(ws, 6, 2, "Emagrecimento")
    cel_amarela(ws, 6, 3, "Estética")
    cel_amarela(ws, 6, 4, "Reposição hormonal")
    ws.cell(row=6, column=5, value="exemplo do cânone da trilha — troque pelas suas").font = \
        Font(name="Calibri", size=9, italic=True, color=CINZA)

    cel_fixa(ws, 7, 1, "Casos nos últimos 30 dias")
    for col, v in ((2, 18), (3, 6), (4, 4)):
        cel_amarela(ws, 7, col, v, fmt='0')

    cel_fixa(ws, 8, 1, "Domínio técnico (nota 1 a 5)")
    for col, v in ((2, 5), (3, 3), (4, 3)):
        cel_amarela(ws, 8, col, v, fmt='0')
    ws.cell(row=8, column=5, value="1 = pouco domínio, 5 = domínio total").font = \
        Font(name="Calibri", size=9, italic=True, color=CINZA)

    cel_fixa(ws, 9, 1, "Resultado repetível (nota 1 a 5)")
    for col, v in ((2, 5), (3, 3), (4, 3)):
        cel_amarela(ws, 9, col, v, fmt='0')
    ws.cell(row=9, column=5, value="1 = varia de paciente pra paciente, 5 = protocolo estável").font = \
        Font(name="Calibri", size=9, italic=True, color=CINZA)

    cel_fixa(ws, 10, 1, "Sustenta plano rentável ao custo da sua hora (nota 1 a 5)")
    for col, v in ((2, 5), (3, 4), (4, 3)):
        cel_amarela(ws, 10, col, v, fmt='0')
    ws.cell(row=10, column=5, value="compare com o custo da sua hora da peça 1 (R$ 408 no exemplo)").font = \
        Font(name="Calibri", size=9, italic=True, color=CINZA)

    secao(ws, 12, "2 · A NOTA QUE APONTA A VENCEDORA", ncols=5)
    cel_fixa(ws, 13, 1, "Nota de volume (a partir dos casos)")
    for col in (2, 3, 4):
        l = get_column_letter(col)
        cel_verde(ws, 13, col, f"=MIN(5,ROUNDUP({l}7/5,0))", fmt='0')
    ws.cell(row=13, column=5, value="1 ponto a cada 5 casos no mês, até o teto de 5").font = \
        Font(name="Calibri", size=9, italic=True, color=CINZA)

    cel_fixa(ws, 14, 1, "NOTA TOTAL", negrito=True, cor=VERDE)
    for col in (2, 3, 4):
        l = get_column_letter(col)
        cel_verde(ws, 14, col, f"={l}13+{l}8+{l}9+{l}10", fmt='0', forte=True)

    cel_fixa(ws, 16, 1, "LINHA VENCEDORA (maior nota total)", negrito=True, cor=VERDE)
    ws.merge_cells("B16:D16")
    cel_verde(ws, 16, 2, "=INDEX(B6:D6,MATCH(MAX(B14:D14),B14:D14,0))", forte=True)

    nota_rodape(ws, 18, "Escolha a vencedora como linha principal pelos próximos 90 dias. "
                        "Dá pra revisar depois de ver os números dos meses seguintes — o "
                        "que não dá é seguir sem nunca ter escolhido.", ncols=5)

    larguras(ws, {"A": 42, "B": 20, "C": 20, "D": 20, "E": 34})
    return wb


# =============================================================== peça 04
def construir_modelo_de_plano():
    """Estrutura do plano de 3 meses. Soma dos minutos / 60 tem que dar 3,5 h
    com o exemplo do cânone."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo de plano"
    ws.sheet_view.showGridLines = False

    cabecalho_sheet(ws, "MODELO DE PLANO DE 3 MESES",
                     "Preencha as células amarelas com o plano da sua linha. As verdes se "
                     "calculam sozinhas. Semana 4 da trilha.")

    secao(ws, 4, "1 · OS BLOCOS DO PLANO")
    cabecalho_tabela(ws, 5, ["Etapa", "Quando", "Duração (min)", "O que o paciente recebe"])
    ws.row_dimensions[5].height = 20

    blocos = [
        (6, "Consulta inicial", "Dia 0", 60,
         "Anamnese completa, histórico, revisão de exames, e o compromisso de "
         "acompanhamento fechado ali mesmo"),
        (7, "Retorno 1", "≈ 3 semanas depois", 30,
         "Ajuste de conduta, primeira resposta ao plano"),
        (8, "Retorno 2", "≈ 7-8 semanas depois", 30,
         "Ajuste de conduta, exame novo se houver"),
        (9, "Retorno 3", "≈ 11 semanas depois — fecha os 3 meses", 30,
         "Fecha o plano e define os próximos passos"),
        (10, "Acompanhamento entre consultas", "Ao longo dos 3 meses", 60,
         "Mensagem de ajuste de dose, retorno de exame fora de consulta, dúvida no "
         "meio da semana"),
    ]
    for linha, etapa, quando, duracao, recebe in blocos:
        cel_fixa(ws, linha, 1, etapa)
        cel_amarela(ws, linha, 2, quando, wrap=True)
        cel_amarela(ws, linha, 3, duracao, fmt='0')
        cel_amarela(ws, linha, 4, recebe, wrap=True)
        ws.row_dimensions[linha].height = 34

    secao(ws, 12, "2 · O TEMPO TOTAL")
    calculado(ws, 13, "Tempo total (minutos)", "=SUM(C6:C10)", fmt='0')
    calculado(ws, 14, "TEMPO TOTAL DO PLANO (horas)", "=B13/60", forte=True, fmt='0.00',
              dica="esse número vira preço na peça 5")

    nota_rodape(ws, 16, "Esse total ainda não é preço — é o tamanho real do plano, em "
                        "horas. É ele que a próxima peça transforma em preço.")

    larguras(ws, {"A": 30, "B": 26, "C": 14, "D": 46, "E": 10})
    return wb


# =============================================================== peça 05
def construir_precificacao():
    """Mesma lógica (inalterada) de `gerar_planilha.py`, já aprovada pelo dono
    do produto — mantida aqui só para a ferramenta continuar regenerável se um
    dia faltar o arquivo já pronto do scratchpad."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Precificação"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "PRECIFICAÇÃO DO SEU PLANO DE ACOMPANHAMENTO"
    t.fill = TITULO_FILL
    t.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    ws["A2"].value = ("Preencha apenas as células AMARELAS. As verdes se calculam sozinhas. "
                      "Semana 5 da trilha.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="4A534E")
    ws.row_dimensions[2].height = 22

    secao(ws, 4, "1 · QUANTO SAI DO CONSULTÓRIO POR MÊS")
    entrada(ws, 5, "Aluguel e condomínio", 4500)
    entrada(ws, 6, "Secretária (com encargos)", 2800)
    entrada(ws, 7, "Prontuário e software", 300)
    entrada(ws, 8, "Contador", 600)
    entrada(ws, 9, "Energia, internet, limpeza", 700)
    entrada(ws, 10, "Marketing", 1200)
    entrada(ws, 11, "Outros", 0)
    entrada(ws, 12, "Seu pró-labore", 25000, dica="quanto VOCÊ precisa tirar por mês")
    calculado(ws, 13, "Total a cobrir por mês", "=SUM(B5:B12)", forte=True)

    secao(ws, 15, "2 · QUANTO CUSTA UMA HORA SUA")
    entrada(ws, 16, "Horas atendidas por semana", 20, fmt='0.0',
            dica="só as horas com paciente na sala")
    entrada(ws, 17, "Semanas por mês", 4.3, fmt='0.0')
    calculado(ws, 18, "Horas por mês", "=B16*B17", fmt='0.0')
    calculado(ws, 19, "CUSTO DA SUA HORA", "=IF(B18=0,0,B13/B18)", forte=True,
              dica="abaixo disso você paga para atender")

    secao(ws, 21, "3 · QUANTO TEMPO O PLANO CONSOME")
    entrada(ws, 22, "Consulta inicial (minutos)", 60, fmt='0')
    entrada(ws, 23, "Quantidade de retornos", 3, fmt='0')
    entrada(ws, 24, "Duração de cada retorno (min)", 30, fmt='0')
    entrada(ws, 25, "Acompanhamento entre consultas (min)", 60, fmt='0',
            dica="mensagem, ajuste de dose, resultado de exame")
    calculado(ws, 26, "Tempo total do plano (horas)", "=(B22+B23*B24+B25)/60", fmt='0.00')
    calculado(ws, 27, "CUSTO DO PLANO — seu piso", "=B26*B19", forte=True,
              dica="abaixo disso é prejuízo, não desconto")

    secao(ws, 29, "4 · O PREÇO")
    entrada(ws, 30, "Margem desejada (%)", 40, fmt=PCT,
            dica="não é seu salário — o pró-labore já entrou lá em cima")
    entrada(ws, 31, "Imposto sobre faturamento (%)", 11, fmt=PCT,
            dica="Simples costuma ficar entre 6% e 15%")
    calculado(ws, 32, "Preço com margem", "=B27*(1+B30/100)")
    calculado(ws, 33, "PREÇO DE TABELA", "=IF(B31>=100,0,B32/(1-B31/100))", forte=True)
    calculado(ws, 34, "Como o paciente ouve (3x de)", "=B33/3",
              dica="parcelou no cartão? desconte a taxa da operadora")

    secao(ws, 36, "5 · O QUE O DESCONTO FAZ COM A SUA SOBRA")
    cabecalhos = ["Desconto", "Você recebe", "Imposto", "Custo do plano", "Sobra no bolso"]
    for i, h in enumerate(cabecalhos, start=1):
        c = ws.cell(row=37, column=i, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color=CINZA)
        c.border = BORDA

    for i, desc in enumerate([0, 5, 10, 15, 20, 25, 30]):
        r = 38 + i
        a = ws.cell(row=r, column=1, value=desc / 100)
        a.number_format = '0%'
        ws.cell(row=r, column=2, value=f"=$B$33*(1-A{r})").number_format = BRL
        ws.cell(row=r, column=3, value=f"=B{r}*$B$31/100").number_format = BRL
        ws.cell(row=r, column=4, value="=$B$27").number_format = BRL
        s = ws.cell(row=r, column=5, value=f"=B{r}-C{r}-D{r}")
        s.number_format = BRL
        s.font = Font(name="Calibri", size=11, bold=True)
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDA

    ws.conditional_formatting.add(
        "E38:E44",
        CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVO_FILL, font=NEGATIVO_FONT))

    ws.merge_cells("A46:E46")
    ws["A46"].value = ("Regra prática: o desconto não sai do preço, sai da sobra. "
                       "Olhe a última coluna antes de conceder qualquer coisa.")
    ws["A46"].font = Font(name="Calibri", size=10, italic=True, color="4A534E")

    larguras(ws, {"A": 38, "B": 16, "C": 15, "D": 16, "E": 17})
    return wb


# =============================================================== peça 07
def construir_roteiro_5_perguntas():
    """As 5 perguntas da peça 7, na ordem, com o que cada uma faz e uma coluna
    em branco pro médico escrever como VAI perguntar com as palavras dele."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Roteiro 5 perguntas"
    ws.sheet_view.showGridLines = False

    cabecalho_sheet(ws, "ROTEIRO DAS 5 PERGUNTAS",
                     "A ordem importa — não pule nenhuma. Preencha a coluna amarela com "
                     "como VOCÊ vai perguntar, com as suas palavras. Semana 7 da trilha.",
                     ncols=4)

    secao(ws, 4, "1 · AS 5 PERGUNTAS, NA ORDEM", ncols=4)
    cabecalho_tabela(ws, 5, ["Nº", "Pergunta (o roteiro padrão)", "O que ela faz",
                              "Como você vai perguntar (suas palavras)"])
    ws.row_dimensions[5].height = 20

    perguntas = [
        (6, 1,
         "Além do que ficou anotado na triagem, o que fez você marcar essa consulta "
         "agora, e não há seis meses?",
         "Busca o motivo real por trás do motivo genérico da ficha."),
        (7, 2,
         "Hoje, isso atrapalha o quê, na prática — o trabalho, o sono, alguma coisa "
         "específica com os filhos?",
         "O paciente nomeia, com as próprias palavras, o custo que já paga por não "
         "ter resolvido."),
        (8, 3,
         "Você já tentou resolver isso antes? O que fez, e o que aconteceu quando "
         "parou?",
         "Evita repetir o que já falhou e mostra que você está ouvindo o caso, não "
         "aplicando protocolo padrão."),
        (9, 4,
         "Se nada mudar a partir de hoje, como você imagina que vai estar daqui a "
         "um ano?",
         "Deixa o paciente nomear o custo de não tratar — sem você precisar "
         "dramatizar."),
        (10, 5,
         "Com tudo isso, o que faz mais sentido pra você agora: um ajuste pontual, "
         "ou resolver isso de fato, com acompanhamento?",
         "Transfere a decisão pro paciente antes de você falar do plano."),
    ]
    for linha, n, pergunta, funcao in perguntas:
        cel_fixa(ws, linha, 1, n)
        ws.cell(row=linha, column=1).alignment = Alignment(horizontal="center")
        cel_fixa(ws, linha, 2, pergunta, wrap=True)
        cel_fixa(ws, linha, 3, funcao, wrap=True)
        cel_amarela(ws, linha, 4, None, wrap=True)
        ws.row_dimensions[linha].height = 56

    secao(ws, 12, "2 · APLIQUE NAS PRÓXIMAS 3 CONSULTAS", ncols=4)
    cabecalho_tabela(ws, 13, ["Consulta", "Data", "Resultado (R = resolver de fato "
                               "/ P = ajuste pontual)"])
    for i, linha in enumerate((14, 15, 16), start=1):
        cel_fixa(ws, linha, 1, i)
        ws.cell(row=linha, column=1).alignment = Alignment(horizontal="center")
        cel_amarela(ws, linha, 2, None)
        cel_amarela(ws, linha, 3, None)

    cel_fixa(ws, 17, 1, "Quantas viraram \"resolver de fato\" (R)", negrito=True, cor=VERDE)
    ws.merge_cells("A17:B17")
    cel_verde(ws, 17, 3, '=COUNTIF(C14:C16,"R")', fmt='0', forte=True)

    nota_rodape(ws, 19, "R = chegou sozinho em \"resolver de fato\". P = só um ajuste "
                        "pontual — também é informação real, não uma objeção pra "
                        "contornar.", ncols=4)

    larguras(ws, {"A": 10, "B": 52, "C": 42, "D": 40})
    return wb


# =============================================================== peça 09
def construir_regua_de_acompanhamento():
    """A régua da peça 9: dia depois da consulta, canal, quem manda, o que a
    mensagem pergunta. Vem preenchida com a régua sugerida + linhas em branco
    pra ajustar."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Régua"
    ws.sheet_view.showGridLines = False

    cabecalho_sheet(ws, "RÉGUA DE ACOMPANHAMENTO — QUEM SOME NO MÊS 2",
                     "A régua sugerida já vem preenchida (tudo amarelo = ajustável). "
                     "Acrescente linhas se quiser. Semana 9 da trilha.",
                     ncols=4)

    secao(ws, 4, "1 · A RÉGUA", ncols=4)
    cabecalho_tabela(ws, 5, ["Quando", "Canal", "Quem manda", "O que a mensagem pergunta"])
    ws.row_dimensions[5].height = 20

    linhas_regua = [
        (6, "D+3", "WhatsApp", "Secretária",
         "Se ele começou o combinado e se sobrou dúvida prática"),
        (7, "D+15", "WhatsApp", "Secretária",
         "Confirma presença e lembra o horário do retorno 1"),
        (8, "D+35", "WhatsApp", "Médico",
         "Uma pergunta pontual sobre o resultado — mesmo sem consulta marcada"),
        (9, "D+55", "WhatsApp", "Secretária", "Confirma presença do retorno 2"),
        (10, "D+75", "WhatsApp", "Secretária",
         "Confirma presença do retorno 3 e fecha o plano"),
    ]
    for linha, quando, canal, quem, pergunta in linhas_regua:
        cel_amarela(ws, linha, 1, quando)
        cel_amarela(ws, linha, 2, canal)
        cel_amarela(ws, linha, 3, quem)
        cel_amarela(ws, linha, 4, pergunta, wrap=True)
        ws.row_dimensions[linha].height = 30
    ws.cell(row=8, column=5, value="o contato que mais importa: vem do médico, "
                                    "não da secretária").font = Font(
        name="Calibri", size=9, italic=True, color=CINZA)

    # linhas em branco pra ele acrescentar contato próprio
    for linha in (11, 12):
        for col in (1, 2, 3, 4):
            cel_amarela(ws, linha, col, None)
        ws.row_dimensions[linha].height = 22

    secao(ws, 14, "2 · DEPOIS DE 60 DIAS", ncols=4)
    entrada(ws, 15, "Quantos pacientes começaram o plano", 0, fmt='0')
    entrada(ws, 16, "Quantos chegaram ao retorno 2", 0, fmt='0')
    calculado(ws, 17, "Chegaram ao retorno 2", "=IF(B15=0,0,B16/B15)", forte=True, fmt=PCT_CALC,
              dica="esse número é o seu ponto de partida pra melhorar a régua")

    nota_rodape(ws, 19, "R$ 612 (43% do custo do plano) já foi gasto no mês 1, antes de "
                        "qualquer garantia de que o paciente continua — é por isso que o "
                        "contato de D+35 importa mais do que parece.", ncols=4)

    larguras(ws, {"A": 30, "B": 16, "C": 18, "D": 52})
    return wb


# =============================================================== peça 12
def construir_painel_mensal():
    """Os 4 indicadores da peça 12, com 1 coluna por mês. Totais e variação
    calculados por fórmula."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Painel mensal"
    ws.sheet_view.showGridLines = False

    NCOLS = 15  # A (indicador) + B..M (12 meses) + N (resumo) + O (nota)
    cabecalho_sheet(ws, "PAINEL DO DONO",
                     "Preencha um número por mês, sempre no mesmo dia. Totais e variação "
                     "calculam sozinhos. Semana 12 da trilha.", ncols=NCOLS)

    secao(ws, 4, "OS QUATRO NÚMEROS DO PAINEL", ncols=NCOLS)
    meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    cabecalho_tabela(ws, 5, ["Indicador"] + meses + ["Resumo do ano", "Nota"])
    ws.row_dimensions[5].height = 20

    # linha 6 · planos ativos
    cel_fixa(ws, 6, 1, "Planos ativos")
    for i in range(12):
        cel_amarela(ws, 6, 2 + i, None, fmt='0')
    cel_verde(ws, 6, 14, "=AVERAGE(B6:M6)", fmt='0.0')
    ws.cell(row=6, column=15, value="média mensal").font = Font(
        name="Calibri", size=9, italic=True, color=CINZA)

    # linha 7 · variação de planos ativos
    cel_fixa(ws, 7, 1, "Variação vs mês anterior", cor=CINZA)
    cel_neutra(ws, 7, 2)
    for col in range(3, 14):
        l, lp = get_column_letter(col), get_column_letter(col - 1)
        cel_verde(ws, 7, col, f'=IF({lp}6=0,"",({l}6-{lp}6)/{lp}6)', fmt=PCT_CALC)

    # linha 8 · entrada do mês
    cel_fixa(ws, 8, 1, "Entrada do mês (R$)")
    for i in range(12):
        cel_amarela(ws, 8, 2 + i, None, fmt=BRL)
    cel_verde(ws, 8, 14, "=SUM(B8:M8)", fmt=BRL, forte=True)
    ws.cell(row=8, column=15, value="total do ano").font = Font(
        name="Calibri", size=9, italic=True, color=CINZA)

    # linha 9 · sobra depois do custo da hora
    cel_fixa(ws, 9, 1, "Sobra depois do custo da hora (R$)")
    for i in range(12):
        cel_amarela(ws, 9, 2 + i, None, fmt=BRL)
    cel_verde(ws, 9, 14, "=SUM(B9:M9)", fmt=BRL, forte=True)
    ws.cell(row=9, column=15, value="total do ano").font = Font(
        name="Calibri", size=9, italic=True, color=CINZA)
    ws.conditional_formatting.add(
        "B9:M9", CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVO_FILL,
                             font=NEGATIVO_FONT))

    # linha 10 · variação da sobra
    cel_fixa(ws, 10, 1, "Variação da sobra vs mês anterior", cor=CINZA)
    cel_neutra(ws, 10, 2)
    for col in range(3, 14):
        l, lp = get_column_letter(col), get_column_letter(col - 1)
        cel_verde(ws, 10, col, f'=IF({lp}9=0,"",({l}9-{lp}9)/{lp}9)', fmt=PCT_CALC)

    # linha 11 · origem dos pacientes novos (texto, sem total)
    cel_fixa(ws, 11, 1, "Origem principal dos pacientes novos")
    for i in range(12):
        cel_amarela(ws, 11, 2 + i, None)
    ws.cell(row=11, column=15, value="indicação, Instagram, Google, outro").font = Font(
        name="Calibri", size=9, italic=True, color=CINZA)

    nota_rodape(ws, 13, "Reserve uma hora, sem paciente na agenda, sempre no mesmo dia do "
                        "mês. Um número isolado não diz muito — comparado com os últimos "
                        "três ou quatro meses, ele denuncia tendência antes de virar crise.",
                ncols=NCOLS)

    larguras(ws, {"A": 34, "N": 15, "O": 26})
    for m in "BCDEFGHIJKLM":
        ws.column_dimensions[m].width = 10
    return wb


# ===================================================================== main
_FERRAMENTAS = {
    "planilha-custo-hora.xlsx": construir_custo_hora,
    "mapa-de-linha.xlsx": construir_mapa_de_linha,
    "modelo-de-plano.xlsx": construir_modelo_de_plano,
    "roteiro-5-perguntas.xlsx": construir_roteiro_5_perguntas,
    "regua-de-acompanhamento.xlsx": construir_regua_de_acompanhamento,
    "painel-mensal.xlsx": construir_painel_mensal,
}


def gerar_tudo(destino=DEST_DIR):
    os.makedirs(destino, exist_ok=True)
    gerados = []

    for nome, construtor in _FERRAMENTAS.items():
        caminho = os.path.join(destino, nome)
        construtor().save(caminho)
        gerados.append(caminho)

    # planilha-precificacao.xlsx: já existe pronta e aprovada — copia o
    # arquivo em vez de regerar. Se a fonte não existir mais (scratchpad é
    # efêmero), cai pro código embutido acima (mesma lógica, mesmo resultado).
    destino_precificacao = os.path.join(destino, "planilha-precificacao.xlsx")
    if os.path.isfile(FONTE_PRECIFICACAO_APROVADA):
        shutil.copy(FONTE_PRECIFICACAO_APROVADA, destino_precificacao)
    else:
        construir_precificacao().save(destino_precificacao)
    gerados.append(destino_precificacao)

    return gerados


if __name__ == "__main__":
    for caminho in gerar_tudo():
        tamanho = os.path.getsize(caminho)
        print(f"gerado: {caminho} ({tamanho} bytes)")
